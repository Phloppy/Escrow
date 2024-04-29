import os
import pandas as pd
import openpyxl
import numpy as np
import re
from fuzzywuzzy import process
from tqdm import tqdm
from pandas import ExcelWriter
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter


# Replace with the path to the folder - omit last backslash '\'
input_folder_path = r'C:\Users\jasonjasinski\OneDrive - At World Properties\Documents\Test\Escrow\Input' + '\\'

# Specify input file names
cibcFile = 'CIBC.xlsx'
lwFile = 'Lonewolf.xlsx'

# Creates full input file paths
cibcPath = os.path.join(input_folder_path, cibcFile)
lwPath = os.path.join(input_folder_path, lwFile)

# Read files at previous paths - sheet name is currently hard coded
cibcinfo = pd.read_excel(cibcPath, sheet_name='cibc')
lonewolf = pd.read_excel(lwPath, sheet_name='lonewolf')

# Specify Output Path
output_folder_path = r'C:\Users\jasonjasinski\OneDrive - At World Properties\Documents\Test\Escrow\Output' + '\\'

# Specify name of output Excel file
output_file = 'escrowRecon.xlsx'

# Concatenate full output path
output_path = os.path.join(output_folder_path, output_file)

# Specify date format
date_style = NamedStyle(name='custom_date', number_format='YYYY/MM/DD')



# Define conditions for categorization
conditions_cibc = [
    ((('BAI Description', ['ACH CREDIT']), ('Detail', ['EARNNEST'])), 'EARNNEST'),
    ((('BAI Description', ['BOOK TRANSFER DEBIT']), ('Detail', ['FUNDS TRANSFER'])), 'FUNDS TRANSFER'),
    ((('BAI Description', ['BOOK TRANSFER CREDIT']), ('Detail', ['FUNDS TRANSFER'])), 'INCOMING FUNDS TRANSFER'),
    ((('BAI Description', ['REMOTE DEPOSIT', 'DEPOSIT ITEM RETURNED']),), 'REMOTE DEPOSIT'),
    ((('BAI Description', ['CHECK PAID']),), 'CHECK PAID'),
    ((('BAI Description', ['INCOMING WIRE TRANSFER', 'OUTGOING WIRE TRANSFER']),), 'WIRE TRANSFER')
]

# Conditions for lonewolf
conditions_lonewolf = [
    (('ref', ['EARNNEST', 'EARNEST', 'EAR', 'NEST']), 'EARNNEST'),
    (('ref', ['WIRE', 'Other_Keywords_As_Needed']), 'WIRE'),
    (('ref', ['EFT', 'Other_Keywords_As_Needed']), 'EFT'),
    (('type', 'ref'), lambda type_val, ref_val: 'C' in type_val and 'EFT' not in ref_val, 'CHECK')
]

# Define a function to create categorization functions based on conditions
def create_categorizer(conditions, default_category="Other"):
    def categorize_row(row):
        for condition in conditions:
            if isinstance(condition[1], tuple):  # It's a keyword condition
                column, keywords = condition[0]
                category = condition[1]
                if any(keyword.upper() in (row[column] if pd.notna(row[column]) else '').upper() for keyword in keywords):
                    return category
            else:  # It's a lambda condition with potentially multiple columns
                columns, lambda_func, category = condition
                if lambda_func(*[row[col] for col in columns]):  # Apply the lambda function to the column values
                    return category
        return default_category
    return categorize_row

# Run categorization function using predefined conditions
categorize_cibc = create_categorizer(conditions_cibc, default_category="Other")
# Apply the categorization function
cibcinfo['Category'] = cibcinfo.apply(categorize_cibc, axis=1)
# Categorization for lonewolf
categorize_lonewolf = create_categorizer(conditions_lonewolf, default_category="REMOTE DEPOSIT")
# Apply the categorization function
lonewolf['Category'] = lonewolf.apply(categorize_lonewolf, axis=1)

#Print confirmation of categorization
print("Categorization Complete")

# Create an ExcelWriter object and use it to write data to separate sheets
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    cibcinfo.to_excel(writer, sheet_name='cibcFull', index=False)
    lonewolf.to_excel(writer, sheet_name='lonewolfFull', index=False)

    # Get the workbook and add the date style
    workbook = writer.book
    workbook.add_named_style(date_style)
    
    # Apply formatting to 'Post' column in 'cibcFull' sheet
    cibc_sheet = writer.sheets['cibcFull']
    # Find the column letter dynamically
    for col_num, cell in enumerate(cibc_sheet[1], 1):
        if cell.value == 'Post':
            post_column = get_column_letter(col_num)
            for cell in cibc_sheet[post_column][1:]:  # Skip header row
                cell.style = date_style
            break
    
    # Apply formatting to 'date' column in 'lonewolfFull' sheet
    lw_sheet = writer.sheets['lonewolfFull']
    # Find the column letter dynamically
    for col_num, cell in enumerate(lw_sheet[1], 1):
        if cell.value == 'date':
            date_column = get_column_letter(col_num)
            for cell in lw_sheet[date_column][1:]:  # Skip header row
                cell.style = date_style
            break


print("Recon Complete")


