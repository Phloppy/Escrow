import os
import pandas as pd
import openpyxl
import numpy as np
import re
from fuzzywuzzy import process
from tqdm import tqdm
from pandas import ExcelWriter
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter

########## INITIAL SETUP ###########

# Replace with the path to the folder - omit last backslash '\'
input_folder_path = r'C:\Users\jasonjasinski\OneDrive - At World Properties\Documents\Test\Escrow\Input' + '\\'

# Specify input file names
cibcFile = 'CIBC.xlsx'
lwFile = 'Lonewolf.xlsx'

# Creates full input file paths
cibcPath = os.path.join(input_folder_path, cibcFile)
lwPath = os.path.join(input_folder_path, lwFile)

# Read files at previous paths - sheet name is currently hard coded
cibcinfo = pd.read_excel(cibcPath, sheet_name='cibc')
lonewolf = pd.read_excel(lwPath, sheet_name='lonewolf')

# Specify Output Folder Path
output_folder_path = r'C:\Users\jasonjasinski\OneDrive - At World Properties\Documents\Test\Escrow\Output' + '\\'

# Specify name of output Excel file
output_file = 'escrowRecon.xlsx'

# Concatenate full output path
output_path = os.path.join(output_folder_path, output_file)

# Specify date format
date_style = NamedStyle(name='custom_date', number_format='YYYY/MM/DD')

########## CATEGORIZATION ###########

### Categorize Lonewolf ###
def categorize_lonewolf(row):
    # Check for 'EARNNEST' and similar keywords in 'ref'
    if any(keyword in row['ref'].upper() for keyword in ['EARNNEST', 'EARNEST', 'EAR', 'NEST']):
        return 'EARNNEST'
    # Check for 'WIRE' and other related keywords in 'ref'
    elif any(keyword in row['ref'].upper() for keyword in ['WIRE', 'Other_Keywords_As_Needed']):
        return 'WIRE'
    # Check for 'EFT' and other related keywords in 'ref'
    elif any(keyword in row['ref'].upper() for keyword in ['EFT', 'Other_Keywords_As_Needed']):
        return 'EFT'
    # Check for 'C/R' and other related keywords in 'ref'
    elif any(keyword in row['ref'].upper() for keyword in ['C/R', 'Other_Keywords_As_Needed']):
        return 'REMOTE DEPOSIT'
    # New condition to check 'type' for 'C' and ensure 'EFT' is not in row['ref']
    elif 'C' in row['type'] and 'EFT' not in row['ref']:
        return 'CHECK'
    # Default category if no conditions are met
    return 'Other'

# Apply the categorization function
lonewolf['Category'] = lonewolf.apply(categorize_lonewolf, axis=1)

### Categorize CIBC ###
def categorize_cibc(row):
    if 'ACH CREDIT' in row['BAI Description'] and 'earnnest' in row['Detail'].lower():
        return 'EARNNEST'
    elif 'BOOK TRANSFER DEBIT' in row['BAI Description'] and 'FUNDS TRANSFER' in row['Detail']:
        return 'FUNDS TRANSFER'
    elif 'BOOK TRANSFER CREDIT' in row['BAI Description'] and 'FUNDS TRANSFER' in row['Detail']:
        return 'INCOMING FUNDS TRANSFER'
    elif any(desc in row['BAI Description'] for desc in ['REMOTE DEPOSIT', 'DEPOSIT ITEM RETURNED']):
        return 'REMOTE DEPOSIT'
    elif 'CHECK PAID' in row['BAI Description']:
        return 'CHECK PAID'
    elif any(wire in row['BAI Description'] for wire in ['INCOMING WIRE TRANSFER', 'OUTGOING WIRE TRANSFER']):
        return 'WIRE TRANSFER'
    return 'Other'

# Apply the categorization function
cibcinfo['Category'] = cibcinfo.apply(categorize_cibc, axis=1)

# Print confirmation of categorization
print("Categorization Complete")

########## GENERATE IDS ###########

# Generate BaseID
cibcinfo['UniqueID'] = cibcinfo['Post'].dt.strftime('%Y%m%d') + '_' + cibcinfo['Transaction Amount'].astype(str) + '_' + cibcinfo['Category']
lonewolf['UniqueID'] = lonewolf['date'].dt.strftime('%Y%m%d') + '_' + lonewolf['amount'].astype(str) + '_' + lonewolf['Category']

# Sequential counter for transactions with the same BaseID
cibcinfo['Counter'] = cibcinfo.groupby('UniqueID').cumcount() + 1
lonewolf['Counter'] = lonewolf.groupby('UniqueID').cumcount() + 1

# Append the counter to the BaseID to ensure uniqueness
cibcinfo['UniqueID'] = cibcinfo['UniqueID'] + '_' + cibcinfo['Counter'].astype(str)
lonewolf['UniqueID'] = lonewolf['UniqueID'] + '_' + lonewolf['Counter'].astype(str)

# Place UniqueID in the first column
cibcinfo = cibcinfo[['UniqueID'] + [col for col in cibcinfo.columns if col != 'UniqueID']]
lonewolf = lonewolf[['UniqueID'] + [col for col in lonewolf.columns if col != 'UniqueID']]

########## CIBC INFORMATION EXTRACTION ###########

def extract_information(row):
    category = row['Category']
    detail = str(row['Detail'])

    # Default values
    address = ''
    name = ''

    if category == 'EARNNEST':
        # Extract address for EARNNEST
        address = detail[49:] if len(detail) > 49 else ''
    elif category == 'WIRE TRANSFER':
        # Extract name for WIRE TRANSFER
        start = detail.find('ORG ')
        end = detail.find(' OBI', start)
        name = detail[start + 4:end].strip() if start != -1 and end != -1 and start < end else ''
        # Extract address for WIRE TRANSFER
        obi_match = re.search(r'\sOBI\s(.+)', detail)
        if obi_match:
            address = obi_match.group(1).strip()
            print(address)  # This print statement is for debugging

    # Return the extracted address and name
    return address, name

# Apply the function and create new columns
cibcinfo[['address', 'Name']] = cibcinfo.apply(extract_information, axis=1, result_type='expand')

# Ensure 'address' column is of string type
cibcinfo['address'] = cibcinfo['address'].astype(str)

# Print the DataFrame to debug
print("Debug Information:")
print(cibcinfo[['Detail', 'address', 'Name']])

########## TRANSACTION MATCH ###########

### EARNNEST ###
# Extract address from 'Detail' starting from the 50th character
cibcinfo['address'] = cibcinfo['Detail'].apply(lambda x: str(x)[49:] if len(str(x)) > 49 else '')

def match_earnest_transactions(cibc_row, lonewolf_df):
    # Filter Lonewolf transactions by the same category and 'EARNNEST'
    lonewolf_earnest = lonewolf_df[lonewolf_df['Category'] == 'EARNNEST']
    
    # First check: same Date, Category, and Amount
    filtered = lonewolf_earnest[(lonewolf_earnest['date'] == cibc_row['Post']) & (lonewolf_earnest['amount'] == cibc_row['Transaction Amount'])]
    if not filtered.empty:
        best_match, score = process.extractOne(cibc_row['address'], filtered['address'].tolist())
        if score >= 80:
            return 'Match', filtered[filtered['address'] == best_match]['UniqueID'].iloc[0]
        # Second check: Ignore Date
        best_match, score = process.extractOne(cibc_row['address'], lonewolf_earnest[lonewolf_earnest['amount'] == cibc_row['Transaction Amount']]['address'].tolist())
        if score >= 80:
            return 'Match - Date Variance', lonewolf_earnest[lonewolf_earnest['address'] == best_match]['UniqueID'].iloc[0]

    # Third check: Ignore Date and Category
    best_match, score = process.extractOne(cibc_row['address'], lonewolf_df[lonewolf_df['amount'] == cibc_row['Transaction Amount']]['address'].tolist())
    if score >= 80:
        return 'Match - Category Variance', lonewolf_df[lonewolf_df['address'] == best_match]['UniqueID'].iloc[0]

    return 'No Match', None

results = cibcinfo[cibcinfo['Category'] == 'EARNNEST'].apply(lambda row: match_earnest_transactions(row, lonewolf), axis=1)
cibcinfo.loc[cibcinfo['Category'] == 'EARNNEST', 'Match Result'] = results.apply(lambda x: x[0])
cibcinfo.loc[cibcinfo['Category'] == 'EARNNEST', 'Lonewolf UniqueID'] = results.apply(lambda x: x[1])

########## WRITE TO FILE ###########

# Create an ExcelWriter object and use it to write data to separate sheets
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    cibcinfo.to_excel(writer, sheet_name='cibcFull', index=False)
    lonewolf.to_excel(writer, sheet_name='lonewolfFull', index=False)

    # Get the workbook and add the date style
    workbook = writer.book
    workbook.add_named_style(date_style)
    
    # Apply formatting to 'Post' column in 'cibcFull' sheet
    cibc_sheet = writer.sheets['cibcFull']
    # Find the column letter dynamically
    for col_num, cell in enumerate(cibc_sheet[1], 1):
        if cell.value == 'Post':
            post_column = get_column_letter(col_num)
            for cell in cibc_sheet[post_column][1:]:  # Skip header row
                cell.style = date_style
            break
    
    # Apply formatting to 'date' column in 'lonewolfFull' sheet
    lw_sheet = writer.sheets['lonewolfFull']
    # Find the column letter dynamically
    for col_num, cell in enumerate(lw_sheet[1], 1):
        if cell.value == 'date':
            date_column = get_column_letter(col_num)
            for cell in lw_sheet[date_column][1:]:  # Skip header row
                cell.style = date_style
            break

print("Recon Complete")
